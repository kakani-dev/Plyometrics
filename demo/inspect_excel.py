import openpyxl
import os

def inspect_file(file_path):
    print(f"Inspecting: {file_path}")
    if not os.path.exists(file_path):
        print("File does not exist.")
        return
    
    wb = openpyxl.load_workbook(file_path, read_only=True)
    print("Sheets:", wb.sheetnames)
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        print(f"\nSheet: {sheet_name}")
        # Print first 5 rows
        rows = list(sheet.iter_rows(max_row=5, values_only=True))
        for r in rows:
            print(r)
    print("-" * 50)

inspect_file("NeuroPi_Grade8_12_Compact_Assessment_AI_Rules.xlsx")
inspect_file("NeuroPi_Grade8_12_Student Development Intelligence Programme 1.xlsx")
