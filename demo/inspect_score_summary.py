import openpyxl

wb = openpyxl.load_workbook("NeuroPi_Grade8_12_Student Development Intelligence Programme 1.xlsx", data_only=True)

if "Score_Summary" in wb.sheetnames:
    sheet = wb["Score_Summary"]
    print("Score_Summary rows:")
    for idx, r in enumerate(sheet.iter_rows(values_only=True), start=1):
        # print row index and first 8 cells
        row_vals = [cell for cell in r]
        if any(v is not None for v in row_vals):
            print(f"Row {idx}: {row_vals[:8]}")
