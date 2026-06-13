import openpyxl

wb = openpyxl.load_workbook("NeuroPi_Grade8_12_Compact_Assessment_AI_Rules.xlsx", data_only=False)

# Let's inspect formulas in Auto_Score_Summary and how it gets its values
if "Auto_Score_Summary" in wb.sheetnames:
    sheet = wb["Auto_Score_Summary"]
    print("Auto_Score_Summary formulas:")
    for r in list(sheet.iter_rows(max_row=10, values_only=False)):
        row_vals = [cell.value for cell in r]
        # print first few cells of the row
        print(row_vals[:9])
        
print("-" * 50)
# Let's inspect the first row of Student_Response_Template to see how Auto_Score_0_100 is calculated
if "Student_Response_Template" in wb.sheetnames:
    sheet = wb["Student_Response_Template"]
    print("Student_Response_Template row 2:")
    for cell in list(sheet.iter_rows(min_row=2, max_row=2, values_only=False))[0]:
        print(cell.coordinate, cell.value)
