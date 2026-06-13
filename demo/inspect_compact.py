import openpyxl
import os

wb = openpyxl.load_workbook("NeuroPi_Grade8_12_Compact_Assessment_AI_Rules.xlsx", read_only=True)
print("Sheets in Compact:", wb.sheetnames)
for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    print(f"\nSheet: {sheet_name}")
    # Print first 5 rows
    rows = list(sheet.iter_rows(max_row=5, values_only=True))
    for r in rows:
        print(r[:10]) # print first 10 columns of first 5 rows
print("-" * 50)
