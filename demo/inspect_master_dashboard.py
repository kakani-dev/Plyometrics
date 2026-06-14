import openpyxl

wb = openpyxl.load_workbook("NeuroPi_Grade8_12_Student Development Intelligence Programme 1.xlsx", data_only=False)

if "Dashboard" in wb.sheetnames:
    sheet = wb["Dashboard"]
    print("Dashboard sheet rows:")
    for r in list(sheet.iter_rows(max_row=25, values_only=False)):
        row_vals = []
        for cell in r:
            val = cell.value
            if hasattr(val, 'text'):
                val = f"Formula: {val.text}"
            row_vals.append(val)
        if any(v is not None for v in row_vals):
            print(row_vals[:8])
