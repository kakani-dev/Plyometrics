import openpyxl
import json
import os

def clean_value(val):
    if val is None:
        return ""
    return str(val).strip()

def parse_master_questions():
    file_path = "NeuroPi_Grade8_12_Student Development Intelligence Programme 1.xlsx"
    if not os.path.exists(file_path):
        print("Master file not found.")
        return []
    
    wb = openpyxl.load_workbook(file_path, data_only=True)
    if "Question_Bank" not in wb.sheetnames:
        print("Question_Bank sheet not found in master file.")
        return []
    
    sheet = wb["Question_Bank"]
    rows = list(sheet.iter_rows(values_only=True))
    
    # Find header row
    header_row_idx = -1
    for i, row in enumerate(rows):
        if any(clean_value(val) == "QID" for val in row):
            header_row_idx = i
            break
            
    if header_row_idx == -1:
        print("Header 'QID' not found in Question_Bank.")
        return []
        
    headers = [clean_value(h) for h in rows[header_row_idx]]
    questions = []
    
    for row in rows[header_row_idx + 1:]:
        # Skip empty rows or header duplicate
        if not row or clean_value(row[0]) == "" or clean_value(row[0]) == "QID":
            continue
        q_data = {}
        for h, val in zip(headers, row):
            if h:
                q_data[h] = clean_value(val)
        questions.append(q_data)
        
    print(f"Parsed {len(questions)} questions from master Question_Bank.")
    return questions

def parse_compact_rules():
    file_path = "NeuroPi_Grade8_12_Compact_Assessment_AI_Rules.xlsx"
    if not os.path.exists(file_path):
        print("Compact file not found.")
        return {}
        
    wb = openpyxl.load_workbook(file_path, data_only=True)
    result = {}
    
    # Parse Metric_Scoring_Rules
    if "Metric_Scoring_Rules" in wb.sheetnames:
        sheet = wb["Metric_Scoring_Rules"]
        rows = list(sheet.iter_rows(values_only=True))
        # Find header
        header_idx = -1
        for i, r in enumerate(rows):
            if any(clean_value(val) == "Metric" for val in r):
                header_idx = i
                break
        if header_idx != -1:
            headers = [clean_value(h) for h in rows[header_idx]]
            rules = []
            for r in rows[header_idx+1:]:
                if not r or clean_value(r[0]) == "" or clean_value(r[0]) == "Metric":
                    continue
                rule_data = {}
                for h, val in zip(headers, r):
                    if h:
                        rule_data[h] = clean_value(val)
                rules.append(rule_data)
            result["Metric_Scoring_Rules"] = rules
            print(f"Parsed {len(rules)} metric scoring rules.")

    # Parse Student_Response_Template
    if "Student_Response_Template" in wb.sheetnames:
        sheet = wb["Student_Response_Template"]
        rows = list(sheet.iter_rows(values_only=True))
        header_idx = -1
        for i, r in enumerate(rows):
            if any(clean_value(val) == "Assessment_QID" for val in r):
                header_idx = i
                break
        if header_idx != -1:
            headers = [clean_value(h) for h in rows[header_idx]]
            template = []
            for r in rows[header_idx+1:]:
                if not r or clean_value(r[1]) == "" or clean_value(r[1]) == "Assessment_QID":
                    continue
                t_data = {}
                for h, val in zip(headers, r):
                    if h:
                        t_data[h] = clean_value(val)
                template.append(t_data)
            result["Student_Response_Template"] = template
            print(f"Parsed {len(template)} template responses.")

    # Parse Permutation_AI_Rules
    if "Permutation_AI_Rules" in wb.sheetnames:
        sheet = wb["Permutation_AI_Rules"]
        rows = list(sheet.iter_rows(values_only=True))
        header_idx = -1
        for i, r in enumerate(rows):
            if any(clean_value(val) == "Profile_Code" for val in r):
                header_idx = i
                break
        if header_idx != -1:
            headers = [clean_value(h) for h in rows[header_idx]]
            permutations = []
            for r in rows[header_idx+1:]:
                if not r or clean_value(r[0]) == "" or clean_value(r[0]) == "Profile_Code":
                    continue
                p_data = {}
                for h, val in zip(headers, r):
                    if h:
                        p_data[h] = clean_value(val)
                permutations.append(p_data)
            result["Permutation_AI_Rules"] = permutations
            print(f"Parsed {len(permutations)} permutations.")

    # Parse Timing_Bands
    if "Timing_Bands" in wb.sheetnames:
        sheet = wb["Timing_Bands"]
        rows = list(sheet.iter_rows(values_only=True))
        header_idx = -1
        for i, r in enumerate(rows):
            if any(clean_value(val) == "Question_Type" for val in r):
                header_idx = i
                break
        if header_idx != -1:
            headers = [clean_value(h) for h in rows[header_idx]]
            timing = []
            for r in rows[header_idx+1:]:
                if not r or clean_value(r[0]) == "" or clean_value(r[0]) == "Question_Type":
                    continue
                t_data = {}
                for h, val in zip(headers, r):
                    if h:
                        t_data[h] = clean_value(val)
                timing.append(t_data)
            result["Timing_Bands"] = timing
            print(f"Parsed {len(timing)} timing bands.")
            
    # Parse Scoring_Map from the master file too since it has interpretation details
    result["Scoring_Map"] = parse_master_scoring_map()

    return result

def parse_master_scoring_map():
    file_path = "NeuroPi_Grade8_12_Student Development Intelligence Programme 1.xlsx"
    if not os.path.exists(file_path):
        return []
    wb = openpyxl.load_workbook(file_path, data_only=True)
    if "Scoring_Map" not in wb.sheetnames:
        return []
    sheet = wb["Scoring_Map"]
    rows = list(sheet.iter_rows(values_only=True))
    header_idx = -1
    for i, r in enumerate(rows):
        if any(clean_value(val) == "Domain" for val in r):
            header_idx = i
            break
    if header_idx == -1:
        return []
    headers = [clean_value(h) for h in rows[header_idx]]
    scoring_map = []
    for r in rows[header_idx+1:]:
        if not r or clean_value(r[0]) == "" or clean_value(r[0]) == "Domain":
            continue
        sm_data = {}
        for h, val in zip(headers, r):
            if h:
                sm_data[h] = clean_value(val)
        scoring_map.append(sm_data)
    print(f"Parsed {len(scoring_map)} scoring map rules.")
    return scoring_map

def main():
    questions = parse_master_questions()
    compact_rules = parse_compact_rules()
    
    # Write files
    with open("questions.json", "w", encoding="utf-8") as f:
        json.dump(questions, f, indent=2, ensure_ascii=False)
        
    with open("rules.json", "w", encoding="utf-8") as f:
        json.dump(compact_rules, f, indent=2, ensure_ascii=False)
        
    print("Done writing JSON files.")

if __name__ == "__main__":
    main()
